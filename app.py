"""
RecrutIA - Application Flask de recrutement intelligent pour ADER Fes
"""

import os, pickle, re, io, secrets, unicodedata
from functools import wraps
from dotenv import load_dotenv
load_dotenv()
import pandas as pd
import pdfplumber
from flask import (Flask, render_template, request, redirect,
                   url_for, flash, jsonify, session)
from flask_sqlalchemy import SQLAlchemy
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta

# ─── CONFIGURATION ─────────────────────────────────────────────────────────────

BASE_DIR = os.path.abspath(os.path.dirname(__file__))

app = Flask(__name__)
app.secret_key = "ader_recrut_ia_2024"

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("rh_logged_in"):
            flash("Accès réservé au personnel RH. Veuillez vous connecter.", "warning")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("rh_logged_in"):
            return redirect(url_for("login"))
        if session.get("role") != "admin":
            flash("Accès réservé aux administrateurs.", "danger")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)
    return decorated
app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{os.path.join(BASE_DIR, 'recrut.db')}"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["UPLOAD_FOLDER"] = os.path.join(BASE_DIR, "static", "uploads")
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024  # 10 MB
os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)

db = SQLAlchemy(app)

ALLOWED_EXTENSIONS = {"pdf"}

# ─── MODÈLE ML ─────────────────────────────────────────────────────────────────

def charger_ou_entrainer_modele():
    model_path   = os.path.join(BASE_DIR, "models", "rf_model.pkl")
    encoder_path = os.path.join(BASE_DIR, "models", "encoders.pkl")

    if os.path.exists(model_path) and os.path.exists(encoder_path):
        with open(model_path, "rb") as f:
            model = pickle.load(f)
        with open(encoder_path, "rb") as f:
            encoders = pickle.load(f)
        return model, encoders

    # Entraînement automatique si les .pkl n'existent pas (ex: Railway)
    from sklearn.ensemble import RandomForestClassifier
    from sklearn.preprocessing import LabelEncoder

    data_path = os.path.join(BASE_DIR, "data", "candidats.csv")
    df = pd.read_csv(data_path, encoding="utf-8-sig")

    niveau_map = {"DOCTORAT":5,"MASTER":4,"LICENCE":3,"TECHNICIEN SPECIALISE":2,"TECHNICIEN":1,"BAC":0}
    df["niveau_diplome"] = df["diplome"].map(niveau_map).fillna(1)
    df["annee_diplome_anciennete"] = datetime.now().year - df["promotion"]

    le_poste      = LabelEncoder()
    le_ecole      = LabelEncoder()
    le_specialite = LabelEncoder()
    df["poste_enc"]      = le_poste.fit_transform(df["poste"].str.upper().str.strip())
    df["ecole_enc"]      = le_ecole.fit_transform(df["ecole"].str.upper().str.strip())
    df["specialite_enc"] = le_specialite.fit_transform(df["specialite"].str.upper().str.strip())

    FEATURES = ["poste_enc","niveau_diplome","ecole_enc","specialite_enc",
                "experience_ans","annee_diplome_anciennete"]
    X, y = df[FEATURES], df["selectionne"]

    model = RandomForestClassifier(n_estimators=200, max_depth=10,
                                   min_samples_split=5, random_state=42,
                                   class_weight="balanced")
    model.fit(X, y)

    encoders = {"le_poste": le_poste, "le_ecole": le_ecole,
                "le_specialite": le_specialite, "features": FEATURES}

    os.makedirs(os.path.join(BASE_DIR, "models"), exist_ok=True)
    with open(model_path, "wb") as f:
        pickle.dump(model, f)
    with open(encoder_path, "wb") as f:
        pickle.dump(encoders, f)

    return model, encoders

RF_MODEL, ENCODERS = charger_ou_entrainer_modele()

DIPLOME_NIVEAU = {
    "DOCTORAT": 5, "MASTER": 4, "LICENCE": 3,
    "TECHNICIEN SPECIALISE": 2, "TECHNICIEN": 1, "BAC": 0,
}

POSTES = ["CGM", "ARCHIVISTE", "SI"]
DIPLOMES = ["MASTER", "LICENCE", "TECHNICIEN SPECIALISE", "TECHNICIEN", "DOCTORAT", "BAC"]

# ─── BASE DE DONNÉES ───────────────────────────────────────────────────────────

class Candidat(db.Model):
    id           = db.Column(db.Integer, primary_key=True)
    nom          = db.Column(db.String(100))
    prenom       = db.Column(db.String(100))
    email        = db.Column(db.String(120))
    telephone    = db.Column(db.String(20))
    poste        = db.Column(db.String(50))
    diplome      = db.Column(db.String(50))
    specialite   = db.Column(db.String(100))
    ecole        = db.Column(db.String(100))
    promotion    = db.Column(db.Integer)
    experience   = db.Column(db.Integer)
    score_ia     = db.Column(db.Float)
    decision     = db.Column(db.String(20))
    decision_manuelle      = db.Column(db.Boolean, default=False)
    decideur_manuel         = db.Column(db.String(150))
    date_decision_manuelle  = db.Column(db.DateTime)
    langues      = db.Column(db.String(200))
    competences  = db.Column(db.Text)
    cv_filename  = db.Column(db.String(200))
    date_depot   = db.Column(db.DateTime, default=datetime.utcnow)
    offre_id     = db.Column(db.Integer, db.ForeignKey('offre.id'), nullable=True)
    offre        = db.relationship('Offre', backref='candidats')

    def to_dict(self):
        return {
            "id": self.id,
            "nom": self.nom,
            "prenom": self.prenom,
            "poste": self.poste,
            "diplome": self.diplome,
            "specialite": self.specialite,
            "experience": self.experience,
            "score_ia": round(self.score_ia * 100, 1) if self.score_ia else 0,
            "decision": self.decision,
            "date_depot": self.date_depot.strftime("%d/%m/%Y") if self.date_depot else "",
        }

class Offre(db.Model):
    id              = db.Column(db.Integer, primary_key=True)
    titre           = db.Column(db.String(200))
    poste           = db.Column(db.String(100))
    nombre_postes   = db.Column(db.Integer, default=1)
    diplome_requis  = db.Column(db.String(100))
    experience_min  = db.Column(db.Integer, default=0)
    specialite      = db.Column(db.String(200))
    langues         = db.Column(db.String(200))
    missions        = db.Column(db.Text)
    competences     = db.Column(db.Text)
    mots_cles       = db.Column(db.String(500))
    date_limite     = db.Column(db.String(50))
    actif           = db.Column(db.Boolean, default=True)
    date_creation   = db.Column(db.DateTime, default=datetime.utcnow)

class Entretien(db.Model):
    id                   = db.Column(db.Integer, primary_key=True)
    candidat_id          = db.Column(db.Integer, db.ForeignKey('candidat.id'), nullable=False)
    candidat             = db.relationship('Candidat', backref=db.backref('entretien', uselist=False))
    date_entretien       = db.Column(db.DateTime, default=datetime.utcnow)
    evaluateur           = db.Column(db.String(150))
    poste_evaluateur     = db.Column(db.String(100))
    note_presentation    = db.Column(db.Float, default=0)
    note_motivation      = db.Column(db.Float, default=0)
    note_competences     = db.Column(db.Float, default=0)
    note_communication   = db.Column(db.Float, default=0)
    note_culture         = db.Column(db.Float, default=0)
    commentaire          = db.Column(db.Text)
    decision_entretien   = db.Column(db.String(30))
    score_entretien      = db.Column(db.Float)


class EvaluationJury(db.Model):
    """Évaluation individuelle par l'un des 4 membres du jury."""
    id                = db.Column(db.Integer, primary_key=True)
    candidat_id       = db.Column(db.Integer, db.ForeignKey('candidat.id'), nullable=False)
    candidat          = db.relationship('Candidat', backref=db.backref('evaluations_jury', lazy=True))
    numero_jury       = db.Column(db.Integer, nullable=False)   # 1, 2, 3 ou 4
    nom_jury          = db.Column(db.String(150))
    poste_jury        = db.Column(db.String(100))
    note_presentation = db.Column(db.Float, default=0)
    note_motivation   = db.Column(db.Float, default=0)
    note_competences  = db.Column(db.Float, default=0)
    note_communication= db.Column(db.Float, default=0)
    note_culture      = db.Column(db.Float, default=0)
    commentaire       = db.Column(db.Text)
    score_jury        = db.Column(db.Float)   # moyenne des 5 critères /20
    date_evaluation   = db.Column(db.DateTime, default=datetime.utcnow)


class VerificationDossier(db.Model):
    """
    Vérification des pièces originales apportées le jour de l'entretien.
    Le candidat passe l'entretien dans tous les cas.
    Si non conforme : éliminé en fin de processus, motif conservé en interne,
    email générique envoyé (sans révéler le vrai motif).
    """
    id                     = db.Column(db.Integer, primary_key=True)
    candidat_id            = db.Column(db.Integer, db.ForeignKey('candidat.id'), nullable=False)
    candidat               = db.relationship('Candidat', backref=db.backref('verification', uselist=False))
    date_verification      = db.Column(db.DateTime, default=datetime.utcnow)
    verificateur           = db.Column(db.String(150))
    conforme               = db.Column(db.Boolean, default=True)
    # Détail par document
    diplome_conforme       = db.Column(db.Boolean, default=True)
    experience_conforme    = db.Column(db.Boolean, default=True)
    cin_conforme           = db.Column(db.Boolean, default=True)
    autres_conformes       = db.Column(db.Boolean, default=True)
    # Motif interne (jamais transmis au candidat)
    motif_interne          = db.Column(db.Text)
    email_envoye           = db.Column(db.Boolean, default=False)


class UtilisateurRH(db.Model):
    __tablename__ = "utilisateur_rh"
    id            = db.Column(db.Integer, primary_key=True)
    nom           = db.Column(db.String(100), nullable=False)
    prenom        = db.Column(db.String(100), nullable=False)
    username      = db.Column(db.String(80), unique=True, nullable=False)
    email         = db.Column(db.String(150))
    password_hash = db.Column(db.String(256), nullable=False)
    role          = db.Column(db.String(20), default="rh")   # admin | rh | jury
    actif         = db.Column(db.Boolean, default=True)
    date_creation = db.Column(db.DateTime, default=datetime.utcnow)
    reset_token       = db.Column(db.String(100))
    reset_token_expire = db.Column(db.DateTime)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)


with app.app_context():
    db.create_all()
    # Créer le compte admin par défaut s'il n'existe pas
    if not UtilisateurRH.query.filter_by(username="admin").first():
        admin = UtilisateurRH(nom="Administrateur", prenom="ADER",
                              username="admin", role="admin")
        admin.set_password("ader2024")
        db.session.add(admin)
        db.session.commit()

# ─── FONCTIONS UTILITAIRES ─────────────────────────────────────────────────────

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def encoder_safe(le, valeur):
    """Encode une valeur inconnue en 0 si pas dans les classes du LabelEncoder."""
    valeur = str(valeur).upper().strip()
    if valeur in le.classes_:
        return le.transform([valeur])[0]
    return 0


def predire(poste, diplome, specialite, ecole, experience, promotion):
    """Lance la prédiction ML et retourne (probabilité, décision)."""
    niveau = DIPLOME_NIVEAU.get(diplome.upper().strip(), 1)
    anciennete = datetime.now().year - int(promotion)

    features = [[
        encoder_safe(ENCODERS["le_poste"], poste),
        niveau,
        encoder_safe(ENCODERS["le_ecole"], ecole),
        encoder_safe(ENCODERS["le_specialite"], specialite),
        int(experience),
        anciennete,
    ]]

    proba = RF_MODEL.predict_proba(features)[0][1]
    if proba >= 0.55:
        decision = "Présélectionné"
    elif proba >= 0.45:
        decision = "À examiner"
    else:
        decision = "Non retenu"
    return round(proba, 4), decision


def couleur_decision(decision):
    """Couleur hexa associée à une décision, pour l'export Excel."""
    return {"Présélectionné": "198754", "À examiner": "E66210"}.get(decision, "dc3545")


def _normaliser_mot(mot):
    """Minuscule, sans accents, sans 's' final simple — pour tolérer pluriels/accents à la comparaison."""
    mot = unicodedata.normalize("NFKD", mot).encode("ascii", "ignore").decode("ascii").lower().strip()
    if mot.endswith("s") and len(mot) > 3:
        mot = mot[:-1]
    return mot


def comparer_competences(competences_candidat, competences_offre):
    """Compare les compétences du CV à celles demandées par l'offre (info, n'influence pas le score).
    Comparaison tolérante mot-par-mot : ignore accents et pluriels simples (ex: "projets" == "projet")."""
    requises = [c.strip() for c in (competences_offre or "").split(",") if c.strip()]
    if not requises:
        return None

    mots_candidat = {
        _normaliser_mot(m) for m in re.findall(r"[a-zA-ZÀ-ÿ']+", competences_candidat or "")
    }

    trouvees, manquantes = [], []
    for c in requises:
        mots_requis = [_normaliser_mot(m) for m in re.findall(r"[a-zA-ZÀ-ÿ']+", c)]
        mots_requis = [m for m in mots_requis if len(m) > 2]
        ratio = sum(1 for m in mots_requis if m in mots_candidat) / len(mots_requis) if mots_requis else 0
        (trouvees if ratio >= 0.6 else manquantes).append(c)

    return {
        "trouvees": trouvees,
        "manquantes": manquantes,
        "total": len(requises),
        "nb_trouvees": len(trouvees),
    }


FEATURE_LABELS = {
    "poste_enc":               "Poste ciblé",
    "niveau_diplome":          "Niveau de diplôme",
    "ecole_enc":               "École / Établissement",
    "specialite_enc":          "Spécialité",
    "experience_ans":          "Années d'expérience",
    "annee_diplome_anciennete": "Ancienneté du diplôme",
}

def expliquer_score(candidat):
    """Retourne les contributions SHAP pour un candidat donné."""
    try:
        import shap
        import numpy as np
        niveau     = DIPLOME_NIVEAU.get(candidat.diplome.upper().strip(), 1)
        anciennete = datetime.now().year - int(candidat.promotion)
        raw = [
            encoder_safe(ENCODERS["le_poste"],      candidat.poste),
            niveau,
            encoder_safe(ENCODERS["le_ecole"],      candidat.ecole),
            encoder_safe(ENCODERS["le_specialite"], candidat.specialite),
            int(candidat.experience),
            anciennete,
        ]
        features      = np.array([raw])
        explainer     = shap.TreeExplainer(RF_MODEL)
        shap_values   = explainer.shap_values(features)
        # shap_values peut être un array 3D (new shap) ou list[2] (old shap)
        if isinstance(shap_values, list):
            contributions = shap_values[1][0]
        else:
            contributions = shap_values[0, :, 1] if shap_values.ndim == 3 else shap_values[0]
        feature_names = ENCODERS.get("features", list(FEATURE_LABELS.keys()))
        result = []
        for i, fname in enumerate(feature_names):
            result.append({
                "label":        FEATURE_LABELS.get(fname, fname),
                "valeur_brute": raw[i],
                "contribution": round(float(contributions[i]), 4),
            })
        result.sort(key=lambda x: abs(x["contribution"]), reverse=True)
        return result
    except Exception as e:
        print(f"SHAP error: {e}")
        return []


def generer_questions(candidat):
    """Génère des questions d'entretien personnalisées via Claude API."""
    try:
        import anthropic as anthropic_sdk
        api_key = os.environ.get("ANTHROPIC_API_KEY")
        if not api_key:
            return None
        client = anthropic_sdk.Anthropic(api_key=api_key)
        prompt = f"""Tu es expert RH pour l'Agence de Développement Régional de Fès-Meknès (ADER).
Génère 10 questions d'entretien personnalisées pour ce candidat.

Poste visé : {candidat.poste}
Diplôme : {candidat.diplome} en {candidat.specialite} ({candidat.ecole}, {candidat.promotion})
Expérience : {candidat.experience} an(s)
Score IA de correspondance : {round(candidat.score_ia * 100, 1)}%

Règles :
- 3 questions techniques sur le poste
- 3 questions sur le parcours et l'expérience
- 2 questions de motivation / projet professionnel
- 2 questions sur la connaissance de l'ADER et du secteur public
- Questions précises, directes, sans jargon inutile
- Format : numérotées 1 à 10, une question par ligne, sans explication"""

        message = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=1024,
            messages=[{"role": "user", "content": prompt}]
        )
        return message.content[0].text
    except Exception:
        return None


def extraire_cv_ia(texte):
    """Extrait les informations d'un CV via l'API Claude. Retourne None si indisponible/échec."""
    try:
        import anthropic as anthropic_sdk
        import json as json_module

        api_key = os.environ.get("ANTHROPIC_API_KEY")
        if not api_key:
            return None

        client = anthropic_sdk.Anthropic(api_key=api_key)
        prompt = f"""Nous sommes en {datetime.now().year}. Voici le texte extrait d'un CV. Analyse-le et réponds UNIQUEMENT avec un objet JSON valide (sans texte avant/après, sans balises markdown), avec exactement ces champs :

- "nom" : nom de famille
- "prenom" : prénom
- "email" : adresse email
- "telephone" : numéro de téléphone
- "diplome" : un seul parmi DOCTORAT, MASTER, LICENCE, TECHNICIEN SPECIALISE, TECHNICIEN, BAC (le plus haut diplôme obtenu). Un diplôme d'Ingénieur / Ingénieur d'État compte comme MASTER (niveau Bac+5 équivalent).
- "specialite" : domaine d'études principal (ex: "Data Science", "Finance", "Génie civil")
- "ecole" : nom de l'établissement du dernier diplôme
- "promotion" : année d'obtention du dernier diplôme (nombre entier)
- "experience" : nombre total d'années d'expérience professionnelle (nombre entier). Pour un poste marqué "en cours"/"présent"/"aujourd'hui", calcule jusqu'à l'année actuelle ({datetime.now().year}), pas une date antérieure.
- "langues" : langues parlées, séparées par des virgules (ex: "Français, Arabe, Anglais")
- "competences" : compétences techniques et mots-clés clés du CV (logiciels, méthodes, domaines maîtrisés), séparés par des virgules (ex: "Excel, gestion de projet, marchés publics")

Si une information est absente du CV, mets une chaîne vide "" (ou 0 pour les nombres).

Texte du CV :
{texte[:6000]}"""

        message = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=500,
            messages=[{"role": "user", "content": prompt}]
        )
        reponse = message.content[0].text.strip()
        reponse = re.sub(r"^```(?:json)?|```$", "", reponse, flags=re.MULTILINE).strip()
        data = json_module.loads(reponse)

        return {
            "nom":        str(data.get("nom", "")).strip().upper(),
            "prenom":     str(data.get("prenom", "")).strip().capitalize(),
            "email":      str(data.get("email", "")).strip(),
            "telephone":  str(data.get("telephone", "")).strip(),
            "diplome":    str(data.get("diplome", "")).strip().upper(),
            "specialite": str(data.get("specialite", "")).strip(),
            "ecole":      str(data.get("ecole", "")).strip(),
            "promotion":  int(data.get("promotion") or 2020),
            "experience": int(data.get("experience") or 0),
            "langues":     str(data.get("langues", "")).strip(),
            "competences": str(data.get("competences", "")).strip(),
        }
    except Exception as e:
        print(f"Erreur extraction CV via IA : {e}")
        return None


def extraire_mots_cles_offre(competences_texte):
    """Condense le paragraphe de compétences d'une offre en mots-clés courts via l'IA."""
    if not competences_texte or not competences_texte.strip():
        return ""
    try:
        import anthropic as anthropic_sdk

        api_key = os.environ.get("ANTHROPIC_API_KEY")
        if not api_key:
            return ""

        client = anthropic_sdk.Anthropic(api_key=api_key)
        prompt = f"""Voici les compétences requises pour un poste, extraites d'une annonce de recrutement. Résume-les en une liste courte de mots-clés (5 à 8 maximum), chacun de 2 à 4 mots, sans phrase complète, sans numérotation, séparés uniquement par des virgules. Réponds UNIQUEMENT avec la liste, rien d'autre.

Texte :
{competences_texte[:3000]}"""

        message = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=200,
            messages=[{"role": "user", "content": prompt}]
        )
        return message.content[0].text.strip()
    except Exception as e:
        print(f"Erreur extraction mots-clés offre : {e}")
        return ""


def extraire_cv(filepath):
    """Extrait les informations clés d'un CV PDF (IA en priorité, règles en secours)."""
    infos = {
        "nom": "", "prenom": "", "email": "", "telephone": "",
        "diplome": "", "specialite": "", "ecole": "",
        "promotion": 2020, "experience": 0,
        "langues": "", "competences": "",
    }
    try:
        with pdfplumber.open(filepath) as pdf:
            texte = "\n".join(
                page.extract_text() or "" for page in pdf.pages
            )

        infos_ia = extraire_cv_ia(texte)
        if infos_ia:
            return infos_ia

        lignes = [l.strip() for l in texte.split("\n") if l.strip()]

        # ── Nom / Prénom : première ligne non vide ──────────────────────────
        if lignes:
            mots = lignes[0].split()
            if len(mots) >= 2:
                infos["prenom"] = mots[0].capitalize()
                infos["nom"] = " ".join(mots[1:]).upper()

        # ── Email ────────────────────────────────────────────────────────────
        m = re.search(r"[\w.+-]+@[\w-]+\.[a-zA-Z]{2,}", texte)
        if m:
            infos["email"] = m.group()

        # ── Téléphone (France +33 ou Maroc +212 ou 06/07) ───────────────────
        m = re.search(
            r"(?:\+33|0033|0)[1-9](?:[\s.\-]?\d{2}){4}"
            r"|(?:\+212|0212|0)[5-7]\d{8}",
            texte
        )
        if m:
            infos["telephone"] = re.sub(r"[\s.\-]", "", m.group())

        # ── Diplôme ──────────────────────────────────────────────────────────
        diplome_map = {
            "doctorat": "DOCTORAT", "phd": "DOCTORAT", "ph.d": "DOCTORAT", "doctorate": "DOCTORAT",
            "master 2": "MASTER", "master2": "MASTER", "mater 2": "MASTER",
            "master 1": "MASTER", "master": "MASTER", "msc": "MASTER",
            "mastère": "MASTER", "mastere": "MASTER",
            "ingénieur d'état": "MASTER", "ingenieur d'etat": "MASTER",
            "diplôme d'ingénieur": "MASTER", "ingénieur": "MASTER", "ingenieur": "MASTER",
            "engineering degree": "MASTER", "bachelor of engineering": "MASTER",
            "master of engineering": "MASTER", "beng": "MASTER", "meng": "MASTER",
            "licence": "LICENCE", "bachelor": "LICENCE", "bachelor's degree": "LICENCE",
            "bachelor of science": "LICENCE", "bsc": "LICENCE",
            "technicien specialise": "TECHNICIEN SPECIALISE",
            "technicien spécialisé": "TECHNICIEN SPECIALISE",
            "technicien": "TECHNICIEN",
            "dut": "TECHNICIEN SPECIALISE",
            "bts": "TECHNICIEN SPECIALISE",
            "associate degree": "TECHNICIEN SPECIALISE",
            "bac": "BAC", "high school diploma": "BAC",
        }
        texte_lower = texte.lower()
        for mot, valeur in diplome_map.items():
            if mot in texte_lower:
                infos["diplome"] = valeur
                break

        # ── Spécialité : mots-clés domaines ─────────────────────────────────
        specialites_map = {
            "intelligence artificielle": "Intelligence Artificielle",
            "data science": "Data Science",
            "informatique": "Informatique",
            "système d'information": "Système d'information",
            "réseaux": "Réseaux et télécommunications",
            "génie logiciel": "Génie logiciel",
            "finance": "Finance",
            "gestion": "Gestion des entreprises",
            "comptabilit": "Comptabilité",
            "marketing": "Marketing",
            "commerce": "Commerce",
            "génie des procédés": "Génie des procédés",
            "génie civil": "Génie civil",
        }
        for mot, valeur in specialites_map.items():
            if mot in texte_lower:
                infos["specialite"] = valeur
                break

        # ── École ────────────────────────────────────────────────────────────
        ecoles_map = {
            "nexa": "NEXA Digital School",
            "encg": "ENCG", "usmba": "USMBA", "ensa": "ENSA",
            "est ": "EST", "ista": "ISTA", "fsjes": "FSJES",
            "claude bernard": "Université Claude Bernard Lyon 1",
            "lyon 1": "Université Claude Bernard Lyon 1",
            "fede": "FEDE", "umr": "UMR", "ektec": "EKTEC",
        }
        for mot, valeur in ecoles_map.items():
            if mot in texte_lower:
                infos["ecole"] = valeur
                break

        # ── Promotion (année du diplôme) ──────────────────────────────────────
        # Priorité : une année juste à côté d'un mot lié au diplôme (fiable),
        # sinon repli sur la dernière année mentionnée dans tout le texte (moins fiable,
        # car ça peut confondre avec une date d'expérience professionnelle récente).
        m = re.search(
            r"(?:diplômé[e]?\s+en|promotion|obtenu[e]?\s+en|dipl[oô]me\s+obtenu\s+en)\s*:?\s*(20[0-2]\d)",
            texte, re.IGNORECASE
        )
        if m:
            infos["promotion"] = int(m.group(1))
        else:
            annees = re.findall(r"\b(20[0-2]\d)\b", texte)
            if annees:
                infos["promotion"] = int(sorted(annees)[-1])

        # ── Expérience totale : calcul depuis les dates d'emploi ─────────────
        # Cherche patterns "Mois AAAA - Mois AAAA" ou "AAAA - AAAA"
        periodes = re.findall(
            r"(\d{4})\s*[-–à]\s*(?:(?:janvier|février|mars|avril|mai|juin|"
            r"juillet|août|septembre|octobre|novembre|décembre)\s+)?(\d{4}|aujourd'hui|présent|actuel)",
            texte, re.IGNORECASE
        )
        annee_courante = datetime.now().year
        total_mois = 0
        for debut, fin in periodes:
            try:
                d = int(debut)
                f = annee_courante if fin.lower() in ["aujourd'hui", "présent", "actuel"] else int(fin)
                if 1990 <= d <= annee_courante and d <= f:
                    total_mois += (f - d) * 12
            except:
                pass
        if total_mois > 0:
            infos["experience"] = min(round(total_mois / 12), 30)
        else:
            # Fallback : cherche "X an(s) d'expérience"
            m = re.search(r"(\d+)\s*an[s]?\s*d.exp", texte, re.IGNORECASE)
            if m:
                infos["experience"] = int(m.group(1))

    except Exception as e:
        print(f"Erreur extraction CV : {e}")

    return infos

# ─── ROUTES ────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    if session.get("rh_logged_in"):
        return redirect(url_for("dashboard"))
    nb_candidatures = Candidat.query.count()
    nb_postes = Offre.query.count()
    return render_template("index.html", nb_candidatures=nb_candidatures, nb_postes=nb_postes)


# ── Candidat : dépôt de candidature ──────────────────────────────────────────

@app.route("/candidature")
def candidature():
    flash("Les candidatures se font via la plateforme emploi public.", "info")
    return redirect(url_for("index"))


# ── Upload CV → extraction automatique ────────────────────────────────────────

@app.route("/extraire_cv", methods=["POST"])
def extraire_cv_route():
    if "cv" not in request.files:
        return jsonify({"error": "Pas de fichier"}), 400

    fichier = request.files["cv"]
    if not allowed_file(fichier.filename):
        return jsonify({"error": "Format non supporté (PDF uniquement)"}), 400

    cv_filename = secure_filename(fichier.filename)
    filepath = os.path.join(app.config["UPLOAD_FOLDER"], cv_filename)
    fichier.save(filepath)

    infos = extraire_cv(filepath)
    infos["cv_filename"] = cv_filename
    return jsonify(infos)


# ── Login / Logout RH ─────────────────────────────────────────────────────────

@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("rh_logged_in"):
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user = UtilisateurRH.query.filter(
            (UtilisateurRH.username == username) | (UtilisateurRH.email == username)
        ).filter_by(actif=True).first()
        if user and user.check_password(password):
            session["rh_logged_in"] = True
            session["user_id"]      = user.id
            session["user_nom"]     = f"{user.prenom} {user.nom}"
            session["role"]         = user.role
            flash(f"Bienvenue, {user.prenom} {user.nom} !", "success")
            return redirect(url_for("dashboard"))
        flash("Identifiants incorrects ou compte désactivé.", "danger")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    flash("Vous êtes déconnecté.", "info")
    return redirect(url_for("login"))


@app.route("/mot-de-passe-oublie", methods=["GET", "POST"])
def mot_de_passe_oublie():
    if request.method == "POST":
        identifiant = request.form.get("identifiant", "").strip()
        user = UtilisateurRH.query.filter(
            (UtilisateurRH.username == identifiant) | (UtilisateurRH.email == identifiant)
        ).filter_by(actif=True).first()

        if user and user.email:
            import smtplib
            from email.mime.text import MIMEText
            from email.mime.multipart import MIMEMultipart

            token = secrets.token_urlsafe(32)
            user.reset_token = token
            user.reset_token_expire = datetime.utcnow() + timedelta(minutes=30)
            db.session.commit()

            SMTP_SERVER = os.environ.get("SMTP_SERVER", "smtp.gmail.com")
            SMTP_PORT   = int(os.environ.get("SMTP_PORT", 587))
            SMTP_USER   = os.environ.get("SMTP_USER", "")
            SMTP_PASS   = os.environ.get("SMTP_PASS", "")

            if SMTP_USER and SMTP_PASS:
                lien = url_for("reinitialiser_mot_de_passe", token=token, _external=True)
                try:
                    msg = MIMEMultipart("alternative")
                    msg["Subject"] = "RecrutIA ADER — Réinitialisation de votre mot de passe"
                    msg["From"]    = SMTP_USER
                    msg["To"]      = user.email

                    corps = f"""
                    <html><body style="font-family:Arial,sans-serif;color:#1A1A1A;">
                    <div style="max-width:600px;margin:auto;border:1px solid #eee;border-radius:8px;overflow:hidden;">
                      <div style="background:#0D5724;padding:20px;text-align:center;">
                        <h2 style="color:#fff;margin:0;">RecrutIA — ADER Fès</h2>
                      </div>
                      <div style="padding:30px;">
                        <p>Bonjour <strong>{user.prenom} {user.nom}</strong>,</p>
                        <p>Une demande de réinitialisation de mot de passe a été effectuée pour votre compte.
                           Cliquez sur le lien ci-dessous pour choisir un nouveau mot de passe
                           (valable 30 minutes) :</p>
                        <p style="text-align:center;margin:28px 0;">
                          <a href="{lien}" style="background:#0D5724;color:#fff;padding:12px 24px;
                             border-radius:6px;text-decoration:none;font-weight:bold;">
                             Réinitialiser mon mot de passe</a>
                        </p>
                        <p>Si vous n'êtes pas à l'origine de cette demande, ignorez simplement cet email.</p>
                        <p>Cordialement,<br><strong>RecrutIA — ADER Fès</strong></p>
                      </div>
                    </div>
                    </body></html>
                    """
                    msg.attach(MIMEText(corps, "html"))
                    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
                        server.starttls()
                        server.login(SMTP_USER, SMTP_PASS)
                        server.sendmail(SMTP_USER, user.email, msg.as_string())
                except Exception:
                    pass

        flash("Si un compte correspond à cet identifiant, un email de réinitialisation vient d'être envoyé.", "info")
        return redirect(url_for("login"))

    return render_template("mot_de_passe_oublie.html")


@app.route("/reinitialiser-mot-de-passe/<token>", methods=["GET", "POST"])
def reinitialiser_mot_de_passe(token):
    user = UtilisateurRH.query.filter_by(reset_token=token).first()
    if not user or not user.reset_token_expire or user.reset_token_expire < datetime.utcnow():
        flash("Ce lien de réinitialisation est invalide ou a expiré.", "danger")
        return redirect(url_for("mot_de_passe_oublie"))

    if request.method == "POST":
        password = request.form.get("password", "")
        confirmation = request.form.get("confirmation", "")
        if len(password) < 8:
            flash("Le mot de passe doit contenir au moins 8 caractères.", "danger")
            return render_template("reinitialiser_mot_de_passe.html", token=token)
        if password != confirmation:
            flash("Les deux mots de passe ne correspondent pas.", "danger")
            return render_template("reinitialiser_mot_de_passe.html", token=token)

        user.set_password(password)
        user.reset_token = None
        user.reset_token_expire = None
        db.session.commit()
        flash("Votre mot de passe a été réinitialisé. Vous pouvez vous connecter.", "success")
        return redirect(url_for("login"))

    return render_template("reinitialiser_mot_de_passe.html", token=token)


# ── Gestion des utilisateurs RH (admin uniquement) ───────────────────────────

@app.route("/admin/utilisateurs")
@admin_required
def gestion_utilisateurs():
    users = UtilisateurRH.query.order_by(UtilisateurRH.date_creation.desc()).all()
    return render_template("admin_utilisateurs.html", users=users)


@app.route("/admin/utilisateurs/nouveau", methods=["GET", "POST"])
@admin_required
def nouvel_utilisateur():
    if request.method == "POST":
        nom      = request.form.get("nom", "").strip()
        prenom   = request.form.get("prenom", "").strip()
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        role     = request.form.get("role", "rh")

        if UtilisateurRH.query.filter_by(username=username).first():
            flash("Ce nom d'utilisateur existe déjà.", "danger")
        elif not nom or not prenom or not username or not password:
            flash("Tous les champs obligatoires doivent être remplis.", "danger")
        else:
            user = UtilisateurRH(nom=nom, prenom=prenom, username=username, role=role)
            user.set_password(password)
            db.session.add(user)
            db.session.commit()
            flash(f"Compte créé pour {prenom} {nom} (identifiant : {username}).", "success")
            return redirect(url_for("gestion_utilisateurs"))
    return render_template("admin_nouvel_utilisateur.html")


@app.route("/admin/utilisateurs/<int:user_id>/toggle", methods=["POST"])
@admin_required
def toggle_utilisateur(user_id):
    user = UtilisateurRH.query.get_or_404(user_id)
    if user.username == "admin":
        flash("Impossible de désactiver le compte administrateur principal.", "danger")
    else:
        user.actif = not user.actif
        db.session.commit()
        flash(f"Compte {user.username} {'activé' if user.actif else 'désactivé'}.", "info")
    return redirect(url_for("gestion_utilisateurs"))


@app.route("/admin/utilisateurs/<int:user_id>/reset-mdp", methods=["POST"])
@admin_required
def reset_mdp(user_id):
    user = UtilisateurRH.query.get_or_404(user_id)
    new_password = request.form.get("new_password", "").strip()
    if not new_password or len(new_password) < 6:
        flash("Le mot de passe doit contenir au moins 6 caractères.", "danger")
    else:
        user.set_password(new_password)
        db.session.commit()
        flash(f"Mot de passe de {user.username} réinitialisé.", "success")
    return redirect(url_for("gestion_utilisateurs"))


# ── Dashboard RH ──────────────────────────────────────────────────────────────

@app.route("/dashboard")
@login_required
def dashboard():
    candidats = Candidat.query.order_by(Candidat.score_ia.desc()).all()

    stats = {
        "total": len(candidats),
        "preselectionnes": sum(1 for c in candidats if c.decision == "Présélectionné"),
        "a_examiner": sum(1 for c in candidats if c.decision == "À examiner"),
        "non_retenus": sum(1 for c in candidats if c.decision == "Non retenu"),
    }
    stats["taux"] = (
        round(stats["preselectionnes"] / stats["total"] * 100, 1)
        if stats["total"] > 0 else 0
    )

    return render_template("dashboard.html",
                           candidats=candidats, stats=stats)


@app.route("/candidat/<int:candidat_id>")
@login_required
def detail_candidat(candidat_id):
    candidat = Candidat.query.get_or_404(candidat_id)
    match_competences = None
    if candidat.offre:
        match_competences = comparer_competences(candidat.competences, candidat.offre.mots_cles)
    return render_template("detail.html",
                           candidat=candidat,
                           score=round(candidat.score_ia * 100, 1),
                           match_competences=match_competences)


def extraire_offre(filepath):
    """Extrait les informations d'une annonce de recrutement PDF."""
    infos = {
        "titre": "", "poste": "", "nombre_postes": 1,
        "diplome_requis": "", "experience_min": 0,
        "specialite": "", "langues": "",
        "missions": "", "competences": "", "date_limite": "",
    }
    try:
        with pdfplumber.open(filepath) as pdf:
            texte = "\n".join(page.extract_text() or "" for page in pdf.pages)

        lignes = [l.strip() for l in texte.split("\n") if l.strip()]
        texte_lower = texte.lower()

        # Titre et poste
        for ligne in lignes[:5]:
            if "poste" in ligne.lower() or "chef" in ligne.lower() or "cadre" in ligne.lower() or "chargé" in ligne.lower():
                infos["titre"] = ligne.replace("Poste pourvu :", "").replace("Poste :", "").strip()
                infos["poste"] = infos["titre"]
                break

        # Nombre de postes
        m = re.search(r"\((\d+)\)", texte)
        if m:
            infos["nombre_postes"] = int(m.group(1))

        # Diplôme requis
        m = re.search(r"[Dd]iplôme\s*:([^\n\-\.]+)", texte)
        if m:
            infos["diplome_requis"] = m.group(1).strip()[:150]

        # Expérience minimum
        m = re.search(r"(\d+)\s*an[s]?\s*(?:d.expérience|minimum|au moins)", texte, re.IGNORECASE)
        if m:
            infos["experience_min"] = int(m.group(1))
        elif "première expérience" in texte_lower or "premier expérience" in texte_lower:
            infos["experience_min"] = 1

        # Langues
        langues_trouvees = []
        if "arabe" in texte_lower: langues_trouvees.append("Arabe")
        if "français" in texte_lower or "francais" in texte_lower: langues_trouvees.append("Français")
        if "anglais" in texte_lower: langues_trouvees.append("Anglais")
        infos["langues"] = ", ".join(langues_trouvees)

        # Missions
        m = re.search(r"[Mm]issions?\s*:(.*?)(?:Principales attributions|Compétences|Conditions|$)", texte, re.DOTALL)
        if m:
            infos["missions"] = m.group(1).strip()[:1000]

        # Compétences
        m = re.search(r"[Cc]ompétences?\s*(?:requises?)?\s*:(.*?)(?:Critères|Dossier|Date|Modalités|$)", texte, re.DOTALL)
        if m:
            infos["competences"] = m.group(1).strip()[:1000]

        # Date limite
        m = re.search(r"(?:dernier délai|date limite)[^\n]*:\s*([^\n]+)", texte, re.IGNORECASE)
        if m:
            infos["date_limite"] = m.group(1).strip()

    except Exception as e:
        print(f"Erreur extraction offre : {e}")

    return infos


# ── Offres RH ─────────────────────────────────────────────────────────────────

@app.route("/offres")
@login_required
def offres():
    liste = Offre.query.order_by(Offre.date_creation.desc()).all()
    return render_template("offres.html", offres=liste)


@app.route("/offre/creer", methods=["GET", "POST"])
@login_required
def creer_offre():
    if request.method == "POST":
        competences_saisies = request.form.get("competences", "").strip()
        offre = Offre(
            titre          = request.form.get("titre", "").strip(),
            poste          = request.form.get("poste", "").strip(),
            nombre_postes  = int(request.form.get("nombre_postes", 1)),
            diplome_requis = request.form.get("diplome_requis", "").strip(),
            experience_min = int(request.form.get("experience_min", 0)),
            specialite     = request.form.get("specialite", "").strip(),
            langues        = request.form.get("langues", "").strip(),
            missions       = request.form.get("missions", "").strip(),
            competences    = competences_saisies,
            mots_cles      = extraire_mots_cles_offre(competences_saisies),
            date_limite    = request.form.get("date_limite", "").strip(),
            actif          = True,
        )
        db.session.add(offre)
        db.session.commit()
        flash("Offre créée avec succès.", "success")
        return redirect(url_for("offres"))

    return render_template("creer_offre.html")


@app.route("/extraire_offre", methods=["POST"])
@login_required
def extraire_offre_route():
    if "offre_pdf" not in request.files:
        return jsonify({"error": "Pas de fichier"}), 400
    fichier = request.files["offre_pdf"]
    if not allowed_file(fichier.filename):
        return jsonify({"error": "PDF uniquement"}), 400
    filepath = os.path.join(app.config["UPLOAD_FOLDER"], secure_filename(fichier.filename))
    fichier.save(filepath)
    infos = extraire_offre(filepath)
    return jsonify(infos)


@app.route("/offre/<int:offre_id>")
@login_required
def detail_offre(offre_id):
    offre = Offre.query.get_or_404(offre_id)
    candidats = Candidat.query.filter_by(offre_id=offre_id).order_by(Candidat.score_ia.desc()).all()
    stats = {
        "total": len(candidats),
        "preselectionnes": sum(1 for c in candidats if c.decision == "Présélectionné"),
        "a_examiner": sum(1 for c in candidats if c.decision == "À examiner"),
        "non_retenus": sum(1 for c in candidats if c.decision == "Non retenu"),
    }
    stats["taux"] = round(stats["preselectionnes"] / stats["total"] * 100, 1) if stats["total"] > 0 else 0
    return render_template("detail_offre.html", offre=offre, candidats=candidats, stats=stats)


@app.route("/offre/<int:offre_id>/importer", methods=["GET", "POST"])
@login_required
def importer_cvs(offre_id):
    offre = Offre.query.get_or_404(offre_id)
    resultats = []

    if request.method == "POST":
        fichiers = request.files.getlist("cvs")
        for fichier in fichiers:
            if not fichier or not allowed_file(fichier.filename):
                continue
            cv_filename = secure_filename(fichier.filename)
            filepath = os.path.join(app.config["UPLOAD_FOLDER"], cv_filename)
            fichier.save(filepath)

            infos = extraire_cv(filepath)
            poste  = offre.poste or "CGM"
            diplome = infos.get("diplome") or "LICENCE"
            specialite = infos.get("specialite") or offre.specialite or "Générale"
            ecole = infos.get("ecole") or "Autre"
            promotion = infos.get("promotion") or 2020
            experience = infos.get("experience") or 0

            # Vérifier doublon : même nom + prénom pour la même offre
            nom_extrait = infos.get("nom") or ""
            prenom_extrait = infos.get("prenom") or ""
            if nom_extrait and prenom_extrait:
                doublon = Candidat.query.filter_by(
                    nom=nom_extrait, prenom=prenom_extrait, offre_id=offre_id
                ).first()
                if doublon:
                    continue

            proba, decision = predire(poste, diplome, specialite, ecole, experience, promotion)

            candidat = Candidat(
                nom        = infos.get("nom") or "—",
                prenom     = infos.get("prenom") or "—",
                email      = infos.get("email") or "",
                telephone  = infos.get("telephone") or "",
                poste      = poste,
                diplome    = diplome,
                specialite = specialite,
                ecole      = ecole,
                promotion  = int(promotion),
                experience = int(experience),
                score_ia   = proba,
                decision   = decision,
                langues    = infos.get("langues") or "",
                competences= infos.get("competences") or "",
                cv_filename= cv_filename,
                offre_id   = offre_id,
            )
            db.session.add(candidat)
            resultats.append({"fichier": cv_filename, "nom": f"{infos.get('prenom','')} {infos.get('nom','')}",
                               "score": round(proba * 100, 1), "decision": decision})

        db.session.commit()
        flash(f"{len(resultats)} CV(s) importé(s) avec succès.", "success")
        return render_template("importer_cvs.html", offre=offre, resultats=resultats)

    return render_template("importer_cvs.html", offre=offre, resultats=[])


@app.route("/offre/<int:offre_id>/supprimer", methods=["POST"])
@login_required
def supprimer_offre(offre_id):
    offre = Offre.query.get_or_404(offre_id)
    # Supprimer les candidats liés
    Candidat.query.filter_by(offre_id=offre_id).delete()
    db.session.delete(offre)
    db.session.commit()
    flash("Offre supprimée.", "info")
    return redirect(url_for("offres"))


@app.route("/api/candidats")
@login_required
def api_candidats():
    candidats = Candidat.query.order_by(Candidat.score_ia.desc()).all()
    return jsonify([c.to_dict() for c in candidats])


# ── Export Excel ──────────────────────────────────────────────────────────────

@app.route("/export/excel")
@login_required
def export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from flask import make_response
    import io

    candidats = Candidat.query.order_by(Candidat.score_ia.desc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Candidatures RecrutIA"

    # En-tête
    headers = ["ID", "Nom", "Prénom", "Email", "Téléphone", "Poste",
               "Diplôme", "Spécialité", "École", "Promotion",
               "Expérience (ans)", "Score IA (%)", "Décision", "Date dépôt"]
    bleu = "1A3A5C"
    or_c = "C8A84B"

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=bleu)
        cell.alignment = Alignment(horizontal="center")

    # Données
    for row, c in enumerate(candidats, 2):
        ws.cell(row=row, column=1, value=c.id)
        ws.cell(row=row, column=2, value=c.nom)
        ws.cell(row=row, column=3, value=c.prenom)
        ws.cell(row=row, column=4, value=c.email)
        ws.cell(row=row, column=5, value=c.telephone)
        ws.cell(row=row, column=6, value=c.poste)
        ws.cell(row=row, column=7, value=c.diplome)
        ws.cell(row=row, column=8, value=c.specialite)
        ws.cell(row=row, column=9, value=c.ecole)
        ws.cell(row=row, column=10, value=c.promotion)
        ws.cell(row=row, column=11, value=c.experience)
        ws.cell(row=row, column=12, value=round(c.score_ia * 100, 1) if c.score_ia else 0)
        dec_cell = ws.cell(row=row, column=13, value=c.decision)
        dec_cell.font = Font(color=couleur_decision(c.decision), bold=True)
        ws.cell(row=row, column=14, value=c.date_depot.strftime("%d/%m/%Y") if c.date_depot else "")

        if row % 2 == 0:
            for col in range(1, 15):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="F0F4F8")

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = make_response(buf.read())
    response.headers["Content-Disposition"] = "attachment; filename=candidatures_recrut_ia.xlsx"
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response


@app.route("/export/excel/offre/<int:offre_id>")
@login_required
def export_excel_offre(offre_id):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from flask import make_response
    import io

    offre = Offre.query.get_or_404(offre_id)
    candidats = Candidat.query.filter_by(offre_id=offre_id).order_by(Candidat.score_ia.desc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = offre.titre[:30]

    headers = ["Classement", "Nom", "Prénom", "Email", "Téléphone",
               "Diplôme", "Spécialité", "École", "Expérience (ans)",
               "Score IA (%)", "Décision", "Date dépôt"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1A3A5C")
        cell.alignment = Alignment(horizontal="center")

    for row, c in enumerate(candidats, 2):
        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=c.nom)
        ws.cell(row=row, column=3, value=c.prenom)
        ws.cell(row=row, column=4, value=c.email)
        ws.cell(row=row, column=5, value=c.telephone)
        ws.cell(row=row, column=6, value=c.diplome)
        ws.cell(row=row, column=7, value=c.specialite)
        ws.cell(row=row, column=8, value=c.ecole)
        ws.cell(row=row, column=9, value=c.experience)
        ws.cell(row=row, column=10, value=round(c.score_ia * 100, 1) if c.score_ia else 0)
        dec_cell = ws.cell(row=row, column=11, value=c.decision)
        dec_cell.font = Font(color=couleur_decision(c.decision), bold=True)
        ws.cell(row=row, column=12, value=c.date_depot.strftime("%d/%m/%Y") if c.date_depot else "")
        if row % 2 == 0:
            for col in range(1, 13):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="F0F4F8")

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = make_response(buf.read())
    nom_fichier = f"classement_{offre.poste}_{offre.id}.xlsx"
    response.headers["Content-Disposition"] = f"attachment; filename={nom_fichier}"
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response


# ── Envoi email aux présélectionnés ───────────────────────────────────────────

@app.route("/offre/<int:offre_id>/notifier", methods=["POST"])
@login_required
def notifier_preselectionnes(offre_id):
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart

    offre = Offre.query.get_or_404(offre_id)
    candidats = Candidat.query.filter_by(offre_id=offre_id, decision="Présélectionné").all()

    SMTP_SERVER = os.environ.get("SMTP_SERVER", "smtp.gmail.com")
    SMTP_PORT   = int(os.environ.get("SMTP_PORT", 587))
    SMTP_USER   = os.environ.get("SMTP_USER", "")
    SMTP_PASS   = os.environ.get("SMTP_PASS", "")

    if not SMTP_USER or not SMTP_PASS:
        flash("Configuration email manquante. Ajoutez SMTP_USER et SMTP_PASS dans les variables d'environnement.", "warning")
        return redirect(url_for("detail_offre", offre_id=offre_id))

    envoyes = 0
    erreurs = 0
    for c in candidats:
        if not c.email:
            continue
        try:
            msg = MIMEMultipart("alternative")
            msg["Subject"] = f"ADER Fès — Présélection pour le poste : {offre.titre}"
            msg["From"]    = SMTP_USER
            msg["To"]      = c.email

            corps = f"""
            <html><body style="font-family:Arial,sans-serif;color:#1A3A5C;">
            <div style="max-width:600px;margin:auto;border:1px solid #eee;border-radius:8px;overflow:hidden;">
              <div style="background:#1A3A5C;padding:20px;text-align:center;">
                <h2 style="color:#C8A84B;margin:0;">ADER Fès — RecrutIA</h2>
              </div>
              <div style="padding:30px;">
                <p>Madame / Monsieur <strong>{c.prenom} {c.nom}</strong>,</p>
                <p>Nous avons le plaisir de vous informer que votre candidature pour le poste de
                   <strong>{offre.titre}</strong> a été retenue lors de la phase de présélection.</p>
                <p>Vous serez contacté(e) prochainement pour la suite du processus de recrutement
                   (date et lieu de l'entretien).</p>
                <p>Veuillez préparer les documents suivants :</p>
                <ul>
                  <li>CV actualisé</li>
                  <li>Copie légalisée de votre CINE</li>
                  <li>Copies légalisées de vos diplômes</li>
                  <li>Attestations d'expérience professionnelle</li>
                </ul>
                <p>Cordialement,<br><strong>Service des Ressources Humaines<br>ADER Fès</strong></p>
              </div>
              <div style="background:#f0f4f8;padding:10px;text-align:center;font-size:12px;color:#888;">
                ADER Fès — Immeuble 21, Rue Mohammed Diouri, 30000 Fès
              </div>
            </div>
            </body></html>
            """
            msg.attach(MIMEText(corps, "html"))
            with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
                server.starttls()
                server.login(SMTP_USER, SMTP_PASS)
                server.sendmail(SMTP_USER, c.email, msg.as_string())
            envoyes += 1
        except Exception as e:
            print(f"Erreur email {c.email}: {e}")
            erreurs += 1

    flash(f"{envoyes} email(s) envoyé(s) avec succès. {erreurs} échec(s).", "success" if envoyes else "warning")
    return redirect(url_for("detail_offre", offre_id=offre_id))


# ── Modifier un candidat ──────────────────────────────────────────────────────

@app.route("/candidat/<int:candidat_id>/modifier", methods=["GET", "POST"])
@login_required
def modifier_candidat(candidat_id):
    candidat = Candidat.query.get_or_404(candidat_id)
    if request.method == "POST":
        candidat.nom        = request.form.get("nom", "").strip()
        candidat.prenom     = request.form.get("prenom", "").strip()
        candidat.email      = request.form.get("email", "").strip()
        candidat.telephone  = request.form.get("telephone", "").strip()
        candidat.diplome    = request.form.get("diplome", "")
        candidat.specialite = request.form.get("specialite", "").strip()
        candidat.ecole      = request.form.get("ecole", "").strip()
        candidat.promotion  = int(request.form.get("promotion", 2020))
        candidat.experience = int(request.form.get("experience", 0))

        proba, decision = predire(candidat.poste, candidat.diplome, candidat.specialite,
                                  candidat.ecole, candidat.experience, candidat.promotion)
        candidat.score_ia = proba
        candidat.decision = decision

        db.session.commit()
        flash("Candidat mis à jour avec succès.", "success")
        return redirect(url_for("detail_candidat", candidat_id=candidat_id))

    return render_template("modifier_candidat.html", candidat=candidat, diplomes=DIPLOMES)




# ── DÉCISION MANUELLE (candidats « À examiner ») ────────────────────────────

@app.route("/candidat/<int:candidat_id>/decision-manuelle", methods=["POST"])
@login_required
def decision_manuelle_route(candidat_id):
    candidat = Candidat.query.get_or_404(candidat_id)

    if candidat.decision != "À examiner":
        flash("Cette action n'est possible que pour un candidat « À examiner ».", "warning")
        return redirect(url_for("detail_candidat", candidat_id=candidat_id))

    action = request.form.get("action")
    if action == "avancer":
        candidat.decision = "Présélectionné"
        flash(f"{candidat.prenom} {candidat.nom} passe en Présélectionné.", "success")
    elif action == "rejeter":
        candidat.decision = "Non retenu"
        flash(f"{candidat.prenom} {candidat.nom} passe en Non retenu.", "info")
    else:
        flash("Action inconnue.", "danger")
        return redirect(url_for("detail_candidat", candidat_id=candidat_id))

    candidat.decision_manuelle      = True
    candidat.decideur_manuel        = session.get("user_nom", "")
    candidat.date_decision_manuelle = datetime.utcnow()
    db.session.commit()
    return redirect(url_for("detail_candidat", candidat_id=candidat_id))


# ── VÉRIFICATION DOSSIER ────────────────────────────────────────────────────

@app.route("/candidat/<int:candidat_id>/verification", methods=["GET", "POST"])
@login_required
def verification_dossier(candidat_id):
    candidat = Candidat.query.get_or_404(candidat_id)

    if request.method == "POST":
        diplome_ok    = request.form.get("diplome_conforme") == "1"
        experience_ok = request.form.get("experience_conforme") == "1"
        cin_ok        = request.form.get("cin_conforme") == "1"
        autres_ok     = request.form.get("autres_conformes") == "1"
        conforme      = diplome_ok and experience_ok and cin_ok and autres_ok

        motif_parts = []
        if not diplome_ok:    motif_parts.append("Diplôme non conforme à la déclaration initiale")
        if not experience_ok: motif_parts.append("Attestations d'expérience non conformes")
        if not cin_ok:        motif_parts.append("Carte d'identité nationale non conforme")
        if not autres_ok:     motif_parts.append("Autres documents non conformes")
        motif = " / ".join(motif_parts) if motif_parts else ""

        verif = candidat.verification
        if not verif:
            verif = VerificationDossier(candidat_id=candidat_id)
            db.session.add(verif)

        verif.verificateur      = request.form.get("verificateur", "").strip()
        verif.conforme          = conforme
        verif.diplome_conforme  = diplome_ok
        verif.experience_conforme = experience_ok
        verif.cin_conforme      = cin_ok
        verif.autres_conformes  = autres_ok
        verif.motif_interne     = motif
        verif.date_verification = datetime.utcnow()
        db.session.commit()

        if conforme:
            flash("Dossier conforme — le candidat peut passer l'entretien.", "success")
        else:
            flash("Dossier non conforme — noté en interne. Le candidat passe quand même l'entretien.", "warning")

        return redirect(url_for("entretien_jury", candidat_id=candidat_id))

    return render_template("verification_dossier.html", candidat=candidat)


# ── ENTRETIEN JURY (vue d'ensemble des 4 jurys) ──────────────────────────────

@app.route("/candidat/<int:candidat_id>/entretien-jury")
@login_required
def entretien_jury(candidat_id):
    candidat = Candidat.query.get_or_404(candidat_id)
    # Récupère les évaluations déjà soumises, indexées par numéro de jury
    evals = {e.numero_jury: e for e in candidat.evaluations_jury}
    # Calcul score moyen si tous les jurys ont évalué
    score_moyen = None
    if len(evals) == 4:
        score_moyen = round(sum(e.score_jury for e in evals.values()) / 4, 2)
    return render_template("entretien_jury.html",
                           candidat=candidat,
                           evals=evals,
                           score_moyen=score_moyen,
                           range4=range(1, 5))


# ── FORMULAIRE D'UN JURY INDIVIDUEL ─────────────────────────────────────────

@app.route("/candidat/<int:candidat_id>/jury/<int:num>", methods=["GET", "POST"])
@login_required
def jury_form(candidat_id, num):
    if num not in (1, 2, 3, 4):
        return redirect(url_for("entretien_jury", candidat_id=candidat_id))

    candidat = Candidat.query.get_or_404(candidat_id)
    # Cherche une évaluation existante pour ce jury
    eval_existante = EvaluationJury.query.filter_by(
        candidat_id=candidat_id, numero_jury=num
    ).first()

    if request.method == "POST":
        notes = {
            "presentation":  float(request.form.get("note_presentation", 0)),
            "motivation":    float(request.form.get("note_motivation", 0)),
            "competences":   float(request.form.get("note_competences", 0)),
            "communication": float(request.form.get("note_communication", 0)),
            "culture":       float(request.form.get("note_culture", 0)),
        }
        score = round(sum(notes.values()) / len(notes), 2)

        if not eval_existante:
            eval_existante = EvaluationJury(candidat_id=candidat_id, numero_jury=num)
            db.session.add(eval_existante)

        eval_existante.nom_jury           = request.form.get("nom_jury", "").strip()
        eval_existante.poste_jury         = request.form.get("poste_jury", "").strip()
        eval_existante.note_presentation  = notes["presentation"]
        eval_existante.note_motivation    = notes["motivation"]
        eval_existante.note_competences   = notes["competences"]
        eval_existante.note_communication = notes["communication"]
        eval_existante.note_culture       = notes["culture"]
        eval_existante.commentaire        = request.form.get("commentaire", "").strip()
        eval_existante.score_jury         = score
        eval_existante.date_evaluation    = datetime.utcnow()
        db.session.commit()

        # Si les 4 jurys ont évalué → calcul et enregistrement du résultat final
        evals = EvaluationJury.query.filter_by(candidat_id=candidat_id).all()
        if len(evals) == 4:
            score_moyen = round(sum(e.score_jury for e in evals) / 4, 2)

            # Décision : vérifier d'abord la conformité du dossier
            verif = candidat.verification
            if verif and not verif.conforme:
                decision = "Éliminé"
            elif score_moyen < 10:
                decision = "Éliminé"
            elif score_moyen >= 12:
                decision = "Retenu"
            else:
                decision = "Éliminé"

            # Mise à jour ou création de l'Entretien agrégé
            entretien = candidat.entretien
            if not entretien:
                entretien = Entretien(candidat_id=candidat_id)
                db.session.add(entretien)
            entretien.score_entretien    = score_moyen
            entretien.decision_entretien = decision
            entretien.date_entretien     = datetime.utcnow()
            db.session.commit()

            flash(f"Les 4 jurys ont évalué — Score moyen : {score_moyen}/20 — Décision : {decision}", "success")
        else:
            flash(f"Évaluation jury #{num} enregistrée (score : {score}/20). {4 - len(evals)} jury(s) restant(s).", "info")

        return redirect(url_for("entretien_jury", candidat_id=candidat_id))

    return render_template("jury_form.html",
                           candidat=candidat,
                           num=num,
                           eval_existante=eval_existante)


# ── LISTE PRIVÉE NON-CONFORMES ────────────────────────────────────────────────

@app.route("/offre/<int:offre_id>/non-conformes")
@login_required
def non_conformes(offre_id):
    offre = Offre.query.get_or_404(offre_id)
    candidats = Candidat.query.filter_by(offre_id=offre_id).all()
    nc_list = []
    for c in candidats:
        if c.verification and not c.verification.conforme:
            nc_list.append(c)
    return render_template("non_conformes.html", offre=offre, candidats=nc_list)


# ── ENVOYER EMAIL NON-CONFORME ────────────────────────────────────────────────

@app.route("/candidat/<int:candidat_id>/envoyer-email-nc", methods=["POST"])
@login_required
def envoyer_email_non_conforme(candidat_id):
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart

    candidat = Candidat.query.get_or_404(candidat_id)
    if not candidat.email:
        flash("Pas d'adresse email pour ce candidat.", "warning")
        return redirect(request.referrer or url_for("dashboard"))

    SMTP_SERVER = os.environ.get("SMTP_SERVER", "sandbox.smtp.mailtrap.io")
    SMTP_PORT   = int(os.environ.get("SMTP_PORT", 587))
    SMTP_USER   = os.environ.get("SMTP_USER", "")
    SMTP_PASS   = os.environ.get("SMTP_PASS", "")

    offre = candidat.offre

    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = f"ADER Fès — Suite de votre candidature : {offre.titre if offre else 'Poste'}"
        msg["From"]    = SMTP_USER
        msg["To"]      = candidat.email

        corps = f"""
        <html><body style="font-family:Arial,sans-serif;color:#1A1A1A;">
        <div style="max-width:600px;margin:auto;border:1px solid #E5E7EB;border-radius:6px;overflow:hidden;">
          <div style="background:#094A1D;padding:20px;text-align:center;">
            <h2 style="color:#fff;margin:0;font-size:1.1rem;">ADER Fès — Service des Ressources Humaines</h2>
          </div>
          <div style="padding:32px;">
            <p>Madame / Monsieur <strong>{candidat.prenom} {candidat.nom}</strong>,</p>
            <p>Nous avons bien reçu votre candidature pour le poste de <strong>{offre.titre if offre else 'le poste en question'}</strong>
               et nous vous remercions de l'intérêt que vous portez à l'Agence de Développement Régional de Fès.</p>
            <p>Après examen approfondi de votre dossier et suite au processus de sélection,
               nous avons le regret de vous informer que votre candidature <strong>n'a pas été retenue</strong>
               pour la suite du processus de recrutement.</p>
            <p>Nous vous encourageons à continuer à suivre nos offres d'emploi futures et restons
               à votre disposition pour tout renseignement complémentaire.</p>
            <p>En vous souhaitant bonne continuation dans vos démarches professionnelles,</p>
            <p>Cordialement,<br>
               <strong>Service des Ressources Humaines<br>ADER Fès</strong></p>
          </div>
          <div style="background:#F8F9FA;padding:12px;text-align:center;font-size:11px;color:#6B7370;">
            ADER Fès — Immeuble 21, Rue Mohammed Diouri, 30000 Fès &nbsp;|&nbsp; contact@ader-fes.ma
          </div>
        </div>
        </body></html>
        """
        msg.attach(MIMEText(corps, "html"))

        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(SMTP_USER, SMTP_PASS)
            server.sendmail(SMTP_USER, candidat.email, msg.as_string())

        if candidat.verification:
            candidat.verification.email_envoye = True
            db.session.commit()

        flash(f"Email envoyé à {candidat.email}", "success")
    except Exception as e:
        flash(f"Erreur envoi email : {e}", "danger")

    return redirect(request.referrer or url_for("dashboard"))


# ── MISE À JOUR CLASSEMENT (prend en compte les 4 jurys + conformité) ────────

@app.route("/offre/<int:offre_id>/classement")
@login_required
def classement_offre(offre_id):
    offre = Offre.query.get_or_404(offre_id)
    candidats = Candidat.query.filter_by(offre_id=offre_id).all()
    resultats = []
    for c in candidats:
        score_ia = round(c.score_ia * 100, 1) if c.score_ia else 0
        score_entretien = None
        score_final = None
        decision_finale = c.decision
        non_conforme = c.verification and not c.verification.conforme

        if c.entretien:
            score_entretien = c.entretien.score_entretien
            if non_conforme:
                decision_finale = "Éliminé (dossier)"
                score_final = score_ia
            else:
                score_final = round(score_ia * 0.4 + (score_entretien / 20 * 100) * 0.6, 1)
                decision_finale = c.entretien.decision_entretien
        else:
            score_final = score_ia
            if non_conforme:
                decision_finale = "Éliminé (dossier)"

        resultats.append({
            "candidat": c,
            "score_ia": score_ia,
            "score_entretien": score_entretien,
            "score_final": score_final,
            "decision_finale": decision_finale,
            "non_conforme": non_conforme,
            "nb_jurys": len(c.evaluations_jury),
        })

    resultats.sort(key=lambda x: x["score_final"], reverse=True)
    return render_template("classement.html", offre=offre, resultats=resultats)


# ── Explication IA (SHAP) ─────────────────────────────────────────────────────

@app.route("/candidat/<int:candidat_id>/explication-ia")
@login_required
def explication_ia(candidat_id):
    candidat      = Candidat.query.get_or_404(candidat_id)
    contributions = expliquer_score(candidat)
    return render_template("explication_ia.html",
                           candidat=candidat,
                           contributions=contributions)


# ── Questions d'entretien IA ──────────────────────────────────────────────────

@app.route("/candidat/<int:candidat_id>/questions-entretien")
@login_required
def questions_entretien(candidat_id):
    candidat   = Candidat.query.get_or_404(candidat_id)
    questions  = generer_questions(candidat)
    has_api_key = bool(os.environ.get("ANTHROPIC_API_KEY"))
    return render_template("questions_entretien.html",
                           candidat=candidat,
                           questions=questions,
                           has_api_key=has_api_key)


@app.route("/debug-shap/<int:candidat_id>")
def debug_shap(candidat_id):
    import traceback
    candidat = Candidat.query.get_or_404(candidat_id)
    try:
        result = expliquer_score(candidat)
        return jsonify({"ok": True, "count": len(result), "data": result})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e), "trace": traceback.format_exc()})


if __name__ == "__main__":
    app.run(debug=True, port=5000, use_reloader=False)
